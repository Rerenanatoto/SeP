import io
import json
import re
from pathlib import Path

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl.chart import ScatterChart, RadarChart, Reference, Series
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="S&P Methodology + SRI", layout="wide")

APP_DIR = Path(__file__).resolve().parent
ASSETS_DIR = APP_DIR / "assets"
DATA_DIR = APP_DIR / "data"
META_COLS = ["country_name", "country_code", "lt_fc_rating"]

# ============================================================
# Helpers gerais
# ============================================================

def normalize_label(text: str) -> str:
    text = str(text).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def slugify(text: str) -> str:
    text = normalize_label(text).lower()
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text


def coerce_numeric(series: pd.Series) -> pd.Series:
    s = series.astype(str).str.strip()
    s = s.replace({
        "N/A": np.nan,
        "N.M.": np.nan,
        "NM": np.nan,
        "None": np.nan,
        "nan": np.nan,
        "": np.nan,
    })
    return pd.to_numeric(s, errors="coerce")


def show_image(path: Path, caption: str | None = None):
    try:
        st.image(str(path), caption=caption, use_container_width=True)
    except TypeError:
        try:
            st.image(str(path), caption=caption, use_column_width=True)
        except TypeError:
            st.image(str(path), caption=caption)


def st_dataframe_compat(df: pd.DataFrame, **kwargs):
    try:
        st.dataframe(df, **kwargs)
    except TypeError:
        st.dataframe(df)


def st_plotly_chart_compat(fig, use_container_width: bool = True):
    try:
        st.plotly_chart(fig, use_container_width=use_container_width)
    except TypeError:
        st.plotly_chart(fig)


def clamp_score(x: float) -> int:
    return int(max(1, min(6, round(x))))


def round_to_half(x: float) -> float:
    y = round(x * 2) / 2
    return max(1.0, min(6.0, y))


def fmt_score(x: float) -> str:
    x = float(x)
    if x.is_integer():
        return f"{int(x)}"
    return f"{x:.1f}"


# ============================================================
# Constantes da metodologia
# ============================================================
GDP_PC_THRESHOLDS = [
    (53600, 1, "Mais que 53.600"),
    (38100, 2, "38.100–53.600"),
    (22600, 3, "22.600–38.100"),
    (7700, 4, "7.700–22.600"),
    (1500, 5, "1.500–7.700"),
    (0, 6, "Abaixo de 1.500"),
]

MEDIAN_GROWTH_BY_INIT = {"1-2": 1.0, "3": 1.7, "4-6": 2.7}

MONETARY_TABLE8A = [
    {"Score": 1, "Exchange-rate regime": "Reserve currency"},
    {"Score": 2, "Exchange-rate regime": "Actively traded or free-floating currency"},
    {"Score": 3, "Exchange-rate regime": "Managed float, crawling peg / crawl-like arrangement, floating with short track record, or intermittent FX intervention"},
    {"Score": 4, "Exchange-rate regime": "Conventional peg or heavy intervention in the foreign exchange market"},
    {"Score": 5, "Exchange-rate regime": "Hard peg (currency board)"},
    {"Score": 6, "Exchange-rate regime": "No local currency (uses another sovereign's currency)"},
]

MONETARY_TABLE8B = {
    1: {
        "monetary_authority_independence": "Strong and long-established track record (more than 10 years) of full independence with clear objectives",
        "monetary_policy_tools_and_effectiveness": "Wide array of monetary instruments",
        "price_stability": "CPI is low and in line with trading partners, leading to stable REER over the economic cycle; broad price stability by other measures",
        "lender_of_last_resort": "Ability to act as lender of last resort for the financial system",
        "local_financial_system_and_capital_markets": "Depository corporation claims on residents in local currency and nonsovereign local-currency bond market capitalization combined exceed 50% of GDP",
    },
    2: {
        "monetary_authority_independence": "Track record of independence",
        "monetary_policy_tools_and_effectiveness": "Market-based monetary instruments",
        "price_stability": "CPI is low and in line with trading partners, leading to fairly stable REER over the economic cycle; broad price stability by other measures",
        "lender_of_last_resort": "Ability to act as lender of last resort for the financial system",
        "local_financial_system_and_capital_markets": "Depository corporation claims on residents in local currency and nonsovereign local-currency bond market capitalization combined exceed 50% of GDP",
    },
    3: {
        "monetary_authority_independence": "Independence, although with a shorter track record or less secure",
        "monetary_policy_tools_and_effectiveness": "Market-based monetary instruments, but heavy reliance on reserve requirements",
        "price_stability": "CPI broadly in line with trading partners over the economic cycle; somewhat volatile REER over the cycle",
        "lender_of_last_resort": "Ability to act as lender of last resort for the financial system",
        "local_financial_system_and_capital_markets": "Depository corporation claims on residents in local currency plus nonsovereign local-currency bond market capitalization and equity market capitalization combined exceed 50% of GDP",
    },
    4: {
        "monetary_authority_independence": "Operational independence, but shorter or less secure than at better assessments",
        "monetary_policy_tools_and_effectiveness": "Market-based monetary instruments, but effectiveness may be untested in a downside scenario",
        "price_stability": "Annual CPI below 10%; somewhat volatile REER over the economic cycle",
        "lender_of_last_resort": "Ability to act as lender of last resort for the financial system",
        "local_financial_system_and_capital_markets": "Depository corporation claims on residents in local currency plus nonsovereign local-currency bond market capitalization and equity market capitalization combined are below 50% of GDP",
    },
    5: {
        "monetary_authority_independence": "Independence is limited by perceived political interference",
        "monetary_policy_tools_and_effectiveness": "Monetary statistics are not viewed as credible",
        "price_stability": "Average CPI typically exceeds 10%; volatile REER over the economic cycle",
        "lender_of_last_resort": "Limited ability to act as lender of last resort for the financial system",
        "local_financial_system_and_capital_markets": "Depository corporation claims on residents in local currency plus nonsovereign local-currency bond market capitalization and equity market capitalization combined are significantly below 50% of GDP",
    },
    6: {
        "monetary_authority_independence": "",
        "monetary_policy_tools_and_effectiveness": "",
        "price_stability": "Average CPI typically exceeds 20%; volatile REER over the economic cycle",
        "lender_of_last_resort": "No ability to act as lender of last resort for the financial system",
        "local_financial_system_and_capital_markets": "Depository corporation claims on residents in local currency plus nonsovereign local-currency bond market capitalization and equity market capitalization combined are significantly below 50% of GDP",
    },
}

MONETARY_TABLE8B_SUMMARY = {
    1: "1 – strongest credibility profile",
    2: "2 – very strong credibility profile",
    3: "3 – strong / intermediate credibility profile",
    4: "4 – adequate but less proven credibility profile",
    5: "5 – weak credibility profile",
    6: "6 – weakest credibility profile",
}

CONTINGENT_TABLE7 = {
    "1-5": {"<=50%": "Limited", "50-100%": "Limited", "100-250%": "Limited", "250-500%": "Limited", ">500%": "Limited/Moderate"},
    "6-7": {"<=50%": "Limited", "50-100%": "Limited", "100-250%": "Limited", "250-500%": "Limited/Moderate", ">500%": "Moderate/High"},
    "8-9": {"<=50%": "Limited", "50-100%": "Limited", "100-250%": "Limited/Moderate", "250-500%": "Moderate/High", ">500%": "High/Very High"},
    "10": {"<=50%": "Limited", "50-100%": "Limited/Moderate", "100-250%": "Moderate/High", "250-500%": "High/Very High", ">500%": "High/Very High"},
}

CONTINGENT_TO_DEBT_ADJ = {
    "Limited": 0,
    "Moderate": 1,
    "High": 2,
    "Very High": 3,
    "Limited/Moderate": None,
    "Moderate/High": None,
    "High/Very High": None,
}

INST_TABLE2 = {
    1: {
        "effectiveness": [
            "Proactive policymaking and a strong track record in managing past economic and financial crisis and delivering economic growth",
            "Ability and willingness to implement reforms to ensure sustainable public finances and economic growth over the long term",
            "Cohesive civil society, as evidenced by high social inclusion, prevalence of civic organizations, degree of social order and capacity of political institutions to respond to societal priorities",
        ],
        "transparency": [
            "Extensive checks and balances between institutions",
            "Unbiased enforcement of contracts and respect for rule of law",
            "Free flow of information throughout society, with open debate of policy decisions",
            "Timely and reliable data and statistical information",
        ],
    },
    2: {
        "effectiveness": [
            "Generally strong, but shorter, track record of policies that deliver sustainable public finances and balanced economic growth consistently over the long term",
            "Weaker ability to implement reforms because of a slow or complex decision-making process",
            "Cohesive civil society, but slightly less in degree than countries we assess '1'",
        ],
        "transparency": [
            "Generally effective checks and balances",
            "Unbiased enforcement of contracts and respect for rule of law",
            "Free flow of information throughout society, with open debate of policy decisions",
            "Timely and reliable data and statistical information",
        ],
    },
    3: {
        "effectiveness": [
            "Generally effective policymaking in recent years, promoting sustainable public finances and balanced economic growth. But policy shifts are possible because of changes in administration or the potential destabilizing influences of underlying socioeconomic or significant long-term fiscal challenges",
            "Cohesive civil society, but less in degree than countries we assess '1' or '2', either because of ethnic, racial, or class tensions or because of higher level of crime",
        ],
        "transparency": [
            "Evolving checks and balances between various institutions",
            "Generally unbiased enforcement of contracts and respect for rule of law",
            "Free flow of information throughout society, but with policy decisions not fully and openly debated",
            "Statistical information that may be less timely than for the higher categories or subject to large revisions",
        ],
    },
    4: {
        "effectiveness": [
            "Policy choices may weaken support for sustainable public finances and balanced economic growth",
            "Reduced predictability of future policy responses because of an uncertain or untested succession process or moderate risk of challenges to political institutions resulting from highly centralized decision-making and parts of the population desiring more political or economic participation",
            "Civil society with ethnic, racial, or class tensions; rising crime rates; and a reduced capacity of political institutions to respond to societal priorities. Low probability, however, of social upheaval",
        ],
        "transparency": [
            "More uncertain checks and balances between institutions, less enforcement of contracts and respect for the rule of law than in above categories",
            "Relatively weak transparency, owing to interference by political institutions in the free dissemination of information, material gaps in data, or reporting delays",
        ],
    },
    5: {
        "effectiveness": [
            "Policy choices likely weaken capability and willingness to maintain sustainable public finances and balanced economic growth, and thus, debt service",
            "High risk of challenges to political institutions, possibly involving domestic conflict, because of demands for more economic or political participation by parts of the population, or significant ethnic or religious challenges to the legitimacy of political institutions",
            "Future policy responses are difficult to predict because of a highly polarized political landscape, highly centralized decision-making or an uncertain or untested succession process",
            "Frayed civil society with difficult ethnic, racial, or class tensions; high crime; and a reduced capacity of political institutions to respond to societal priorities. Rising chance of social upheaval",
        ],
        "transparency": [
            "Unassured enforcement of contracts and respect for rule of law",
            "Impaired transparency, owing to at least one of the following factors: moderate to high levels of perceived corruption, material data gaps, or significant interference by political institutions in the free dissemination of information",
        ],
    },
    6: {
        "effectiveness": [
            "Weak political institutions, resulting in an uncertain policy environment in periods of stress, including diminished capability and willingness to maintain timely debt service",
            "Considerable risk of breakdown between political institutions, including significant risk of domestic conflict",
            "Distressed civil society; sharp ethnic, racial, or class tensions; inability or unwillingness of political institutions to respond to societal priorities; or present danger of social upheaval",
        ],
        "transparency": [
            "Unassured enforcement of contracts and respect for rule of law",
            "Impaired transparency, owing to several of the following factors: frequent and material data revisions or lack or suppression of data and information flows; or high levels of perceived corruption of political institutions",
        ],
    },
}

INST_TABLE2_LABELS = {
    1: "1 – Proactive policymaking / extensive checks & balances",
    2: "2 – Strong track record (shorter) / effective checks & balances",
    3: "3 – Effective policymaking (recent) / evolving checks & balances",
    4: "4 – Reduced predictability / weaker transparency",
    5: "5 – High challenges / impaired transparency",
    6: "6 – Weak institutions / impaired transparency",
}

INDICATIVE_MATRIX_COLS = [1.0, 1.5, 2.0, 2.5, 3.0, 3.5, 4.0, 4.5, 5.0, 5.5, 6.0]
INDICATIVE_MATRIX = [
    ["AAA", "AAA", "AAA", "AA+", "AA", "A+", "A", "A-", "BBB+", "BB+", "BB-"],
    ["AAA", "AAA", "AA+", "AA", "AA-", "A", "A-", "BBB+", "BBB", "BB+", "BB-"],
    ["AAA", "AA+", "AA", "AA-", "A", "A-", "BBB+", "BBB", "BB+", "BB", "B+"],
    ["AA+", "AA", "AA-", "A+", "A-", "BBB", "BBB-", "BB+", "BB", "BB-", "B+"],
    ["AA", "AA-", "A+", "A", "BBB+", "BBB-", "BB+", "BB", "BB-", "B+", "B"],
    ["AA-", "A+", "A", "BBB+", "BBB", "BB+", "BB", "BB-", "B+", "B", "B"],
    ["A", "A-", "BBB+", "BBB", "BB+", "BB", "BB-", "B+", "B", "B-", "B-"],
    ["BBB", "BBB", "BBB-", "BB+", "BB", "BB-", "B+", "B", "B", "B-", "B-"],
    ["BB+", "BB+", "BB", "BB-", "B+", "B", "B", "B-", "B-", "B-", "B-"],
]

RATING_SCALE = [
    "AAA", "AA+", "AA", "AA-", "A+", "A", "A-",
    "BBB+", "BBB", "BBB-",
    "BB+", "BB", "BB-",
    "B+", "B", "B-",
    "CCC+", "CCC", "CCC-", "CC", "C",
]

# ============================================================
# Lógica metodologia
# ============================================================

def fp_bucket_index(fp_profile: float) -> int:
    x = fp_profile
    if x <= 1.7: return 0
    if x <= 2.2: return 1
    if x <= 2.7: return 2
    if x <= 3.2: return 3
    if x <= 3.7: return 4
    if x <= 4.2: return 5
    if x <= 4.7: return 6
    if x <= 5.2: return 7
    return 8


def indicative_from_matrix(ie_profile: float, fp_profile: float) -> str:
    col_val = round_to_half(ie_profile)
    try:
        col = INDICATIVE_MATRIX_COLS.index(col_val)
    except ValueError:
        col = min(range(len(INDICATIVE_MATRIX_COLS)), key=lambda i: abs(INDICATIVE_MATRIX_COLS[i] - col_val))
    row = fp_bucket_index(fp_profile)
    return INDICATIVE_MATRIX[row][col]


def apply_notches(base_rating: str, notch_adj: int, lc_uplift: int = 0) -> str:
    r = (base_rating or "").upper().strip()
    if r not in RATING_SCALE:
        return r
    idx = RATING_SCALE.index(r)
    new_idx = idx - int(notch_adj) - int(lc_uplift)
    new_idx = max(0, min(len(RATING_SCALE) - 1, new_idx))
    return RATING_SCALE[new_idx]


def init_economic_from_gdppc(gdppc_usd: float) -> int:
    for thr, score, _label in GDP_PC_THRESHOLDS:
        if gdppc_usd >= thr:
            return score
    return 6


def pick_growth_bucket(init_score: int) -> str:
    if init_score in (1, 2):
        return "1-2"
    if init_score == 3:
        return "3"
    return "4-6"


def table5_candidates(change_net_debt_gdp: float) -> list[int]:
    x = float(change_net_debt_gdp)
    if x < 0: return [1]
    if x < 1: return [1, 2]
    if x < 2: return [2]
    if x < 3: return [2, 3]
    if x < 4: return [3, 4]
    if x < 5: return [4, 5]
    if x < 6: return [5]
    if x < 7: return [5, 6]
    return [6]


def table5_initial_from_inputs(change_net_debt_gdp: float, overlap_trend: str) -> int:
    candidates = table5_candidates(change_net_debt_gdp)
    if len(candidates) == 1:
        return candidates[0]
    if overlap_trend == "melhorando":
        return min(candidates)
    return max(candidates)


def table6_initial_from_inputs(net_debt_gdp: float, interest_to_rev: float) -> int:
    if net_debt_gdp <= 30: col = 0
    elif net_debt_gdp <= 60: col = 1
    elif net_debt_gdp <= 80: col = 2
    elif net_debt_gdp <= 100: col = 3
    else: col = 4
    if interest_to_rev <= 5: row = 0
    elif interest_to_rev <= 10: row = 1
    elif interest_to_rev <= 15: row = 2
    else: row = 3
    matrix = [
        [1, 2, 3, 4, 5],
        [2, 3, 4, 5, 6],
        [3, 4, 5, 6, 6],
        [4, 5, 6, 6, 6],
    ]
    return matrix[row][col]


def radar(scores: dict):
    cats = list(scores.keys())
    vals = [scores[c] for c in cats]
    cats2 = cats + [cats[0]]
    vals2 = vals + [vals[0]]
    DB = "#1F3864"
    fig = go.Figure(data=[go.Scatterpolar(
        r=vals2, theta=cats2,
        fill="toself",
        name="Scores",
        line=dict(color=DB, width=2),
        fillcolor="rgba(31,56,100,0.15)",
        marker=dict(color=DB, size=6),
    )])
    fig.update_layout(
        polar=dict(
            radialaxis=dict(visible=True, range=[6, 1]),
            angularaxis=dict(tickfont=dict(size=12, color=DB)),
        ),
        showlegend=False,
        margin=dict(l=60, r=60, t=40, b=40),
    )
    return fig


def bullets(items):
    return "".join([f"- {x}" for x in items])


def download_payload():
    payload = {
        "institutional": st.session_state.get("institutional", 4),
        "economic": st.session_state.get("economic", 4),
        "external": st.session_state.get("external", 3),
        "fiscal": st.session_state.get("fiscal", 4.0),
        "monetary": st.session_state.get("monetary", 3),
        "profiles": st.session_state.get("profiles", {}),
        "indicative": st.session_state.get("indicative", None),
        "final_rating": st.session_state.get("final_rating", None),
        "notch_adj": st.session_state.get("notch_adj", 0),
        "lc_uplift": st.session_state.get("lc_uplift", 0),
        "notes": st.session_state.get("notes", ""),
    }
    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")


# ============================================================
# Helpers: exporta .xlsx com gráficos nativos openpyxl
# ============================================================

def _sanitize_sheet_name(name: str, existing: list, max_len: int = 28) -> str:
    """Sanitiza nome de aba Excel; evita duplicatas."""
    safe = re.sub(r'[/\\?*\[\]:\']+', "_", str(name)).strip()[:max_len]
    if not safe:
        safe = "Aux"
    candidate, suffix = safe, 2
    while candidate in existing:
        candidate = f"{safe[:max_len-2]}_{suffix}"
        suffix += 1
    return candidate


def _build_sri_scatter_charts(wb, df_s, sheet_name: str):
    """
    Cria um ScatterChart (lineMarker) por indicador.
    - Anos no eixo X via scaling.min/max + majorUnit (API openpyxl).
    - Sem linhas de grade.
    - Aba auxiliar com nome do indicador.
    """
    cols = list(df_s.columns)
    if not all(c in cols for c in ["indicator", "country_name", "year_num", "value"]):
        return
    indicators = sorted(df_s["indicator"].dropna().unique().tolist())
    if not indicators:
        return

    chart_ws_name = "Graficos"
    if chart_ws_name in wb.sheetnames:
        chart_ws_name = f"Graf_{sheet_name[:10]}"
    chart_ws = wb.create_sheet(chart_ws_name)
    existing = list(wb.sheetnames)
    chart_row = 1

    for ind in indicators:
        ind_df = df_s[df_s["indicator"] == ind].dropna(subset=["year_num", "value"])
        if ind_df.empty:
            continue
        countries    = sorted(ind_df["country_name"].dropna().unique().tolist())
        years_sorted = sorted(ind_df["year_num"].dropna().unique().tolist())
        n_years      = len(years_sorted)
        min_yr       = float(years_sorted[0])
        max_yr       = float(years_sorted[-1])

        # Aba auxiliar: nome do indicador (sanitizado)
        aux_name = _sanitize_sheet_name(ind, existing)
        existing.append(aux_name)
        aux = wb.create_sheet(aux_name)
        aux.cell(row=1, column=1, value="Ano")
        for ci, country in enumerate(countries, start=2):
            aux.cell(row=1, column=ci, value=country[:28])
        for ri, yr in enumerate(years_sorted, start=2):
            aux.cell(row=ri, column=1, value=int(yr))
            for ci, country in enumerate(countries, start=2):
                sv = ind_df[
                    (ind_df["year_num"] == yr) & (ind_df["country_name"] == country)
                ]["value"]
                aux.cell(row=ri, column=ci,
                         value=round(float(sv.mean()), 4) if not sv.empty else None)

        sc = ScatterChart()
        sc.scatterStyle = "lineMarker"
        sc.style        = 10
        sc.title        = ind[:50]
        sc.y_axis.title = "Valor"
        sc.x_axis.title = "Ano"
        sc.width        = 22
        sc.height       = 12

        # Forçar anos no eixo X — definir ANTES de add_chart
        sc.x_axis.numFmt          = "0"
        sc.x_axis.tickLblPos      = "low"
        sc.x_axis.crosses         = "min"
        sc.x_axis.orientation     = "minMax"
        sc.x_axis.scaling.min     = min_yr
        sc.x_axis.scaling.max     = max_yr
        sc.x_axis.majorUnit       = 1.0
        sc.x_axis.tickMarkSkip    = 1
        sc.x_axis.tickLblSkip     = 1
        # Remover grades
        sc.x_axis.majorGridlines  = None
        sc.x_axis.minorGridlines  = None
        sc.y_axis.majorGridlines  = None
        sc.y_axis.minorGridlines  = None

        x_ref = Reference(aux, min_col=1, min_row=2, max_row=n_years + 1)
        for ci in range(2, len(countries) + 2):
            y_ref  = Reference(aux, min_col=ci, min_row=2, max_row=n_years + 1)
            series = Series(y_ref, xvalues=x_ref,
                            title=aux.cell(row=1, column=ci).value)
            series.smooth        = False
            series.marker.symbol = "circle"
            series.marker.size   = 4
            sc.series.append(series)

        chart_ws.add_chart(sc, f"A{chart_row}")
        chart_row += 23


def to_excel_bytes(sheets_dict: dict, add_charts: bool = False) -> bytes:
    """
    Converte {nome_aba: DataFrame} em bytes .xlsx.
    add_charts=True:
      - SRI-Dashboards / SRI-Dados → ScatterCharts por indicador
      - Metodologia → RadarChart azul escuro, pilares nos vértices, sem legenda
    """
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for sheet_name, df_s in sheets_dict.items():
            df_s.to_excel(writer, sheet_name=sheet_name[:31], index=False)

        if add_charts:
            wb = writer.book
            for sheet_name, df_s in sheets_dict.items():
                safe = sheet_name[:31]
                cols = list(df_s.columns)

                # ── METODOLOGIA ───────────────────────────────────────────────
                if safe == "Metodologia" and "Parametro" in cols and "Valor" in cols:
                    ws = wb[safe]
                    pillar_params = ["Institutional", "Economic", "External",
                                     "Fiscal", "Monetary"]
                    param_col = df_s["Parametro"].tolist()
                    pillar_rows = []
                    for pname in pillar_params:
                        for i, p in enumerate(param_col):
                            if str(p).strip() == pname:
                                pillar_rows.append(i + 2)
                                break

                    if len(pillar_rows) == 5:
                        # Colunas auxiliares D (pilar) e E (score)
                        ws["D1"] = "Pilar"
                        ws["E1"] = "Score"
                        for ri, (pname, rexcel) in enumerate(
                                zip(pillar_params, pillar_rows), start=2):
                            ws[f"D{ri}"] = pname
                            try:
                                ws[f"E{ri}"] = float(ws[f"B{rexcel}"].value)
                            except (TypeError, ValueError):
                                ws[f"E{ri}"] = 0

                        # RadarChart: linhas sem preenchimento
                        radar = RadarChart()
                        radar.type   = "standard"
                        radar.style  = 10
                        radar.title  = "Perfil de Scores - S&P Metodologia"
                        radar.legend = None        # sem legenda

                        # Escala 0-6, números em cada anel
                        radar.y_axis.delete      = False
                        radar.y_axis.numFmt      = "0"
                        radar.y_axis.scaling.min = 0
                        radar.y_axis.scaling.max = 6
                        radar.y_axis.majorUnit   = 1

                        # Dados e categorias (pilares nos vértices)
                        data_ref = Reference(ws, min_col=5, min_row=1, max_row=6)
                        cats_ref = Reference(ws, min_col=4, min_row=2, max_row=6)
                        radar.add_data(data_ref, titles_from_data=True)
                        radar.set_categories(cats_ref)

                        # Cor azul escuro na série
                        if radar.series:
                            s = radar.series[0]
                            s.graphicalProperties.line.solidFill        = "1F3864"
                            s.graphicalProperties.line.width            = 20000
                            s.marker.symbol                              = "circle"
                            s.marker.size                                = 5
                            s.marker.graphicalProperties.solidFill      = "1F3864"
                            s.marker.graphicalProperties.line.solidFill = "1F3864"

                        radar.width  = 16
                        radar.height = 12
                        ws.add_chart(radar, "D8")

                # ── SRI ───────────────────────────────────────────────────────
                elif safe in ("SRI-Dashboards", "SRI-Dados") and all(
                    c in cols for c in ["indicator", "country_name", "year_num", "value"]
                ):
                    _build_sri_scatter_charts(wb, df_s, safe)

    buf.seek(0)
    return buf.read()


# ============================================================
# Lógica SRI
# ============================================================

def find_local_xlsx() -> Path | None:
    preferred_names = [
        DATA_DIR / "base.xlsx",
        DATA_DIR / "report.xlsx",
        APP_DIR / "base.xlsx",
        APP_DIR / "report.xlsx",
    ]
    for candidate in preferred_names:
        if candidate.exists():
            return candidate
    for search_dir in [DATA_DIR, APP_DIR]:
        if search_dir.exists():
            files = sorted([p for p in search_dir.glob("*.xlsx") if not p.name.startswith("~$")])
            if files:
                return files[0]
    return None


def find_data_end(raw: pd.DataFrame) -> int:
    first_col = raw.iloc[:, 0].astype(str).fillna("").str.strip().str.lower()
    end_idx = len(raw)
    for idx, value in enumerate(first_col):
        if (
            value.startswith("lt fc--")
            or value.startswith("copyright")
            or value.startswith("no content")
            or value.startswith("credit-related")
            or value.startswith("to reprint")
            or value.startswith("any passwords/user ids")
        ):
            end_idx = idx
            break
    return end_idx


def parse_sheet(raw: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
    raw = raw.copy().dropna(how="all", axis=1)
    if len(raw) < 6:
        return pd.DataFrame()
    end_idx = find_data_end(raw)
    raw = raw.iloc[:end_idx].reset_index(drop=True)
    if len(raw) < 6:
        return pd.DataFrame()
    indicator_row = raw.iloc[3].tolist()
    year_row = raw.iloc[4].tolist()
    records = []
    current_indicator = None
    for col_idx, cell in enumerate(indicator_row):
        if col_idx < 3:
            continue
        if pd.notna(cell):
            current_indicator = normalize_label(cell)
        year_value = year_row[col_idx] if col_idx < len(year_row) else None
        if current_indicator and pd.notna(year_value):
            records.append({
                "col_idx": col_idx,
                "indicator": current_indicator,
                "indicator_key": slugify(current_indicator),
                "year": str(year_value).strip(),
            })
    if not records:
        return pd.DataFrame()
    col_map = pd.DataFrame(records)
    data = raw.iloc[5:].copy().dropna(how="all")
    if data.empty:
        return pd.DataFrame()
    rename_map = {}
    if 0 in data.columns: rename_map[0] = "country_name"
    if 1 in data.columns: rename_map[1] = "country_code"
    if 2 in data.columns: rename_map[2] = "lt_fc_rating"
    data = data.rename(columns=rename_map)
    required_cols = ["country_name", "country_code", "lt_fc_rating"]
    if not all(col in data.columns for col in required_cols):
        return pd.DataFrame()
    usable_cols = [c for c in col_map["col_idx"].tolist() if c in data.columns]
    if not usable_cols:
        return pd.DataFrame()
    col_map = col_map[col_map["col_idx"].isin(usable_cols)].copy()
    data = data[required_cols + usable_cols].copy()
    data["country_name"] = data["country_name"].astype(str).str.strip()
    data["country_code"] = data["country_code"].astype(str).str.strip()
    data["lt_fc_rating"] = data["lt_fc_rating"].astype(str).str.strip()
    invalid_starts = ("lt fc--", "copyright", "no content")
    data = data[
        data["country_name"].ne("")
        & ~data["country_name"].str.lower().str.startswith(invalid_starts)
    ].copy()
    if data.empty:
        return pd.DataFrame()
    long_df = data.melt(
        id_vars=required_cols,
        value_vars=usable_cols,
        var_name="col_idx",
        value_name="value_raw",
    )
    long_df = long_df.merge(col_map, on="col_idx", how="left")
    long_df["sheet"] = sheet_name
    long_df["sheet_key"] = slugify(sheet_name)
    long_df["value"] = coerce_numeric(long_df["value_raw"])
    long_df["year"] = long_df["year"].astype(str).str.strip()
    long_df["year_num"] = pd.to_numeric(long_df["year"].str.extract(r"(\d{4})")[0], errors="coerce")
    long_df["is_forecast"] = long_df["year"].str.contains(r"[ef]$", case=False, na=False)
    return long_df[[
        "sheet", "sheet_key", "country_name", "country_code", "lt_fc_rating",
        "indicator", "indicator_key", "year", "year_num", "is_forecast", "value"
    ]]


@st.cache_data(show_spinner=False)
def load_workbook(file_bytes=None) -> pd.DataFrame:
    empty = pd.DataFrame(columns=[
        "sheet", "sheet_key", "country_name", "country_code", "lt_fc_rating",
        "indicator", "indicator_key", "year", "year_num", "is_forecast", "value"
    ])
    if file_bytes is not None:
        source = io.BytesIO(file_bytes)
    else:
        local_file = find_local_xlsx()
        if local_file is None:
            return empty
        source = local_file
    xls = pd.ExcelFile(source, engine="openpyxl")
    frames = []
    for sheet in xls.sheet_names:
        raw = pd.read_excel(xls, sheet_name=sheet, header=None, engine="openpyxl")
        try:
            parsed = parse_sheet(raw, sheet)
            if not parsed.empty:
                frames.append(parsed)
        except Exception:
            continue
    if not frames:
        return empty
    df = pd.concat(frames, ignore_index=True)
    return df.dropna(subset=["indicator", "year"])


def build_filters(df: pd.DataFrame) -> pd.DataFrame:
    st.subheader("Filtros do SRI")
    all_categories = sorted(df["sheet"].dropna().unique().tolist())
    selected_categories = st.multiselect(
        "Categoria",
        options=all_categories,
        default=all_categories,
        key="f_categories",
        help="Escolha uma ou mais categorias (abas da planilha).",
    )
    df1 = df[df["sheet"].isin(selected_categories)] if selected_categories else df.copy()

    all_ratings = sorted(df1["lt_fc_rating"].dropna().unique().tolist())
    selected_ratings = st.multiselect(
        "LT FC rating",
        options=all_ratings,
        default=[],
        key="f_ratings",
        help="Deixe vazio para considerar todos os ratings.",
    )
    df2 = df1[df1["lt_fc_rating"].isin(selected_ratings)] if selected_ratings else df1.copy()

    all_countries = sorted(df2["country_name"].dropna().unique().tolist())
    selected_countries = st.multiselect(
        "País",
        options=all_countries,
        default=[],
        key="f_countries",
        help="Deixe vazio para considerar todos os países do recorte atual.",
    )

    all_indicators = sorted(df2["indicator"].dropna().unique().tolist())
    selected_indicators = st.multiselect(
        "Indicadores",
        options=all_indicators,
        default=[],
        key="f_indicators",
        help="Deixe vazio para considerar todos os indicadores do recorte atual.",
    )

    valid_years = df2["year_num"].dropna()
    year_min, year_max = (2019, 2028) if valid_years.empty else (int(valid_years.min()), int(valid_years.max()))
    selected_year_range = st.slider(
        "Faixa de anos",
        min_value=year_min,
        max_value=year_max,
        value=(year_min, year_max),
        key="f_years",
    )

    forecast_mode = st.radio(
        "Período",
        ["Todos", "Somente históricos", "Somente estimativas/projeções"],
        index=0,
        key="f_forecast",
        horizontal=True,
    )

    filtered = df2.copy()
    if selected_countries:
        filtered = filtered[filtered["country_name"].isin(selected_countries)]
    if selected_indicators:
        filtered = filtered[filtered["indicator"].isin(selected_indicators)]
    filtered = filtered[filtered["year_num"].between(selected_year_range[0], selected_year_range[1], inclusive="both")]
    if forecast_mode == "Somente históricos":
        filtered = filtered[~filtered["is_forecast"]]
    elif forecast_mode == "Somente estimativas/projeções":
        filtered = filtered[filtered["is_forecast"]]
    return filtered


def render_dashboard_tab(df: pd.DataFrame):
    st.subheader("Dashboards")
    if df.empty:
        st.warning("Nenhum dado encontrado com os filtros selecionados.")
        return

    c1, c2, c3 = st.columns(3)
    c1.metric("Países", df["country_name"].nunique())
    c2.metric("Indicadores", df["indicator"].nunique())
    c3.metric("Observações", f"{len(df):,}".replace(",", "."))

    available_sheets = sorted(df["sheet"].dropna().unique().tolist())
    if not available_sheets:
        st.info("Nenhuma aba disponível para plotagem com os filtros atuais.")
        return

    if len(available_sheets) == 1:
        sheet_for_charts = available_sheets[0]
        st.caption(f"Mostrando todos os indicadores da aba: **{sheet_for_charts}**")
    else:
        sheet_for_charts = st.selectbox(
            "Aba para gerar gráficos (um gráfico por indicador)",
            available_sheets,
            index=0,
            key="sheet_for_charts",
        )

    plot_df = df[df["sheet"] == sheet_for_charts].copy().dropna(subset=["year_num", "value"])
    if plot_df.empty:
        st.info("Sem dados numéricos para gerar gráficos nesta aba com os filtros atuais.")
        return

    indicators = sorted(plot_df["indicator"].dropna().unique().tolist())
    if not indicators:
        st.info("Nenhum indicador encontrado para esta aba.")
        return

    st.markdown(f"### {sheet_for_charts} — gráficos para **{len(indicators)}** indicadores")
    cols_per_row = 2
    show_legend = plot_df["country_name"].nunique() <= 12

    for i in range(0, len(indicators), cols_per_row):
        row_inds = indicators[i : i + cols_per_row]
        row_cols = st.columns(cols_per_row)
        for col, ind in zip(row_cols, row_inds):
            with col:
                ind_df = plot_df[plot_df["indicator"] == ind].sort_values(["country_name", "year_num"])
                if ind_df.empty:
                    st.caption(f"Sem dados para: {ind}")
                    continue
                fig = px.line(
                    ind_df,
                    x="year_num",
                    y="value",
                    color="country_name",
                    markers=True,
                    hover_data=["lt_fc_rating", "year", "country_code"],
                    title=ind,
                )
                fig.update_layout(
                    height=340,
                    margin=dict(l=10, r=10, t=50, b=10),
                    legend_title_text="País",
                    showlegend=show_legend,
                )
                fig.update_xaxes(title="Ano")
                fig.update_yaxes(title="Valor")
                st_plotly_chart_compat(fig, use_container_width=True)

    st.markdown("#### Média por LT FC rating (no recorte atual)")
    rating_summary = (
        df.groupby(["lt_fc_rating", "indicator"], as_index=False)["value"]
        .mean()
        .rename(columns={"value": "media_valor"})
        .sort_values(["indicator", "lt_fc_rating"])
    )
    st_dataframe_compat(rating_summary, use_container_width=True, hide_index=True)


    # ── Download Excel ───────────────────────────────────────────────
    st.markdown("---")
    _dash_df = plot_df.reset_index(drop=True) if not plot_df.empty else df.reset_index(drop=True)
    _excel_dash = to_excel_bytes({"SRI-Dashboards": _dash_df}, add_charts=True)
    st.download_button(label="⬇️ Baixar Dashboards (.xlsx com gráficos)",
        data=_excel_dash, file_name="sri_dashboards.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_dashboard_xlsx")


def render_table_tab(df: pd.DataFrame):
    st.subheader("Dados em tabela")
    if df.empty:
        st.warning("Nenhum dado encontrado com os filtros selecionados.")
        return
    view_mode = st.radio(
        "Visualização",
        ["Longa (recomendada)", "Pivotada"],
        horizontal=True,
        index=0,
        key="view_mode",
    )
    if view_mode == "Longa (recomendada)":
        display_df = df.sort_values(["sheet", "country_name", "indicator", "year_num"]).copy()
    else:
        display_df = (
            df.pivot_table(
                index=["sheet", "country_name", "country_code", "lt_fc_rating", "indicator"],
                columns="year",
                values="value",
                aggfunc="first",
            )
            .reset_index()
        )
    st_dataframe_compat(display_df, use_container_width=True, hide_index=True)
    col_dl1, col_dl2 = st.columns(2)
    with col_dl1:
        csv_data = display_df.to_csv(index=False).encode("utf-8-sig")
        st.download_button("⬇️ Baixar CSV", data=csv_data,
            file_name="bda_filtrado.csv", mime="text/csv", key="dl_table_csv")
    with col_dl2:
        _excel_table = to_excel_bytes({"SRI-Dados": display_df}, add_charts=True)
        st.download_button(label="⬇️ Baixar Excel (.xlsx com gráficos)",
            data=_excel_table, file_name="sri_dados.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_table_xlsx")

# ============================================================
# UI metodologia (agora dentro da 1ª aba)
# ============================================================

def render_methodology_tab():
    st.header("Metodologia")
    _metod_data = {
        "Parametro": ["Institutional","Economic","External","Fiscal","Monetary",
                      "IE Profile","FP Profile","Indicative Rating",
                      "Notch Adj.","LC Uplift","Final Rating"],
        "Valor": [
            st.session_state.get("institutional","---"),
            st.session_state.get("economic","---"),
            st.session_state.get("external","---"),
            st.session_state.get("fiscal","---"),
            st.session_state.get("monetary","---"),
            (st.session_state.get("profiles") or {}).get("Institutional & Economic profile","---"),
            (st.session_state.get("profiles") or {}).get("Flexibility & Performance profile","---"),
            st.session_state.get("indicative","---"),
            st.session_state.get("notch_adj","---"),
            st.session_state.get("lc_uplift","---"),
            st.session_state.get("final_rating","---"),
        ],
    }
    _df_metod = pd.DataFrame(_metod_data)
    _excel_metod = to_excel_bytes({"Metodologia": _df_metod}, add_charts=True)
    st.download_button(label="⬇️ Baixar Metodologia (.xlsx com gráfico radar)",
        data=_excel_metod, file_name="metodologia_sp.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_metod_xlsx")
    st.markdown("---")
    method_page = st.radio(
        "Seção da metodologia",
        ["Visão geral", "Economic", "Fiscal", "Monetary", "External", "Institutional", "Resultados"],
        horizontal=True,
        key="method_page_select",
    )
    st.markdown("---")

    if method_page == "Visão geral":
        st.title("📊 S&P Sovereign Rating Methodology – Dashboard")
        st.write(
            "Este dashboard é uma implementação **independente de Excel**, baseada no PDF de critérios. "
            "A metodologia avalia **cinco pilares** (1=mais forte, 6=mais fraco) e combina em dois perfis: "
            "(i) **Institutional & Economic profile** e (ii) **Flexibility & Performance profile**. "
            "A partir disso, consulta-se a matriz de **Indicative Rating Levels** (Tabela 1) e aplica-se julgamento (±1 notch) e eventuais fatores suplementares."
        )
        col1, col2 = st.columns([1, 1], gap="large")
        with col1:
            st.subheader("Estrutura (5 pilares)")
            img = ASSETS_DIR / "page_03_img_01.png"
            if img.exists():
                show_image(img)
            else:
                st.info("Imagem do framework não encontrada em assets/ (opcional).")
        with col2:
            st.subheader("Matriz de Indicative Rating Levels (Tabela 1)")
            img = ASSETS_DIR / "page_06_img_01.png"
            if img.exists():
                show_image(img)
            else:
                st.info("Imagem da matriz não encontrada em assets/ (opcional).")
        st.markdown("---")
        st.subheader("Como usar")
        st.markdown(
            "1) Vá em cada assessment e preencha os inputs."
            "2) Vá em **Resultados** para ver os perfis e um radar com os 5 pilares."
            "3) O app calcula automaticamente o **Indicative rating level** pela Tabela 1, e você aplica notches/uplift."
        )
    elif method_page == "Economic":
        st.title("Economic assessment")
    
        st.markdown("#### 1) Income level (GDP per capita) → score inicial")
        gdppc = st.number_input(
            "GDP per capita (US$) – ano corrente",
            min_value=0.0,
            value=10000.0,
            step=100.0,
            key="eco_gdppc",
        )
    
        init_score = init_economic_from_gdppc(gdppc)
    
        df_thr = pd.DataFrame(
            [
                {"Faixa (US$)": label, "Initial economic assessment": score}
                for _, score, label in GDP_PC_THRESHOLDS
            ]
        )
        st_dataframe_compat(df_thr, use_container_width=True, hide_index=True)
        st.metric("Score inicial (Income level)", init_score)
    
        st.markdown("#### 2) Ajustes (até ±2 categorias no total)")
    
        bucket = pick_growth_bucket(init_score)
        median_growth = MEDIAN_GROWTH_BY_INIT[bucket]
    
        st.caption(
            f"Referência: mediana de crescimento real per capita (10 anos) para init={bucket}: {median_growth:.1f}%"
        )
    
        trend = st.number_input(
            "Real GDP per capita trend growth (10y, %) – sua estimativa",
            value=median_growth,
            step=0.1,
            key="eco_trend",
        )
    
        # ============================================================
        # Ajuste AUTOMÁTICO por growth prospects
        # Regra:
        # - bucket 1-2: melhora se >= mediana + 0.7 ; piora se < mediana - 0.7
        # - bucket 3:   melhora se >= mediana + 0.7 ; piora se < mediana - 0.7
        # - bucket 4-6: melhora se >= mediana + 0.7 ; piora se < mediana
        # Ou seja, para bucket 4-6, abaixo de 2.7 já piora 1 categoria.
        # ============================================================
        if bucket == "4-6":
            if trend >= median_growth + 0.7:
                adj_growth_auto = -1
            elif trend < median_growth:
                adj_growth_auto = 1
            else:
                adj_growth_auto = 0
        else:
            if trend >= median_growth + 0.7:
                adj_growth_auto = -1
            elif trend < (median_growth - 0.7):
                adj_growth_auto = 1
            else:
                adj_growth_auto = 0
    
        c_auto1, c_auto2 = st.columns([1, 2])
        with c_auto1:
            st.metric("Ajuste automático por growth prospects", f"{adj_growth_auto:+d}")
        with c_auto2:
            st.caption(
                "Para bucket 4-6, valores abaixo da mediana (2.7%) já geram piora de 1 categoria."
            )
    
        use_manual_growth_override = st.checkbox(
            "Override manual do ajuste por growth prospects",
            value=False,
            key="eco_use_manual_growth_override",
        )
    
        if use_manual_growth_override:
            adj_growth = st.selectbox(
                "Ajuste manual por growth prospects",
                options=[-1, 0, 1],
                index=[-1, 0, 1].index(adj_growth_auto),
                key="eco_adj_growth_manual",
                help="-1 melhora 1 categoria; +1 piora 1 categoria.",
            )
        else:
            adj_growth = adj_growth_auto
    
        adj_concentration = st.selectbox(
            "Ajuste por concentração/volatilidade (0 ou +1)",
            options=[0, 1],
            index=0,
            key="eco_adj_conc",
        )
    
        adj_credit_bubble = st.selectbox(
            "Ajuste por credit-bubble risk (0 ou +1)",
            options=[0, 1],
            index=0,
            key="eco_adj_bubble",
        )
    
        adj_data = st.selectbox(
            "Ajuste por inconsistência de dados (0 ou +1)",
            options=[0, 1],
            index=0,
            key="eco_adj_data",
        )
    
        total_adj = adj_growth + adj_concentration + adj_credit_bubble + adj_data
        total_adj = max(-2, min(2, total_adj))
    
        final_score = clamp_score(init_score + total_adj)
    
        st.markdown("---")
        st.metric("Economic assessment (final)", final_score)
        st.caption(
            f"Growth prospects: {adj_growth:+d} | "
            f"Concentração/volatilidade: {adj_concentration:+d} | "
            f"Credit-bubble risk: {adj_credit_bubble:+d} | "
            f"Inconsistência de dados: {adj_data:+d}"
        )
        st.caption(f"Total de ajuste aplicado (limitado a ±2): {total_adj:+d}")
    
        st.session_state["economic"] = int(final_score)

    elif method_page == "Fiscal":
        st.title("Fiscal assessment")
        st.caption("Cálculo baseado nas Table 5 e 6, com critérios e ajustes no mesmo estilo da aba External (sem barras/sliders para score inicial).")
        st.markdown("## 1) Fiscal performance & flexibility (Table 5)")
        c1, c2 = st.columns([1, 1])
        with c1:
            change_net_debt_gdp = st.number_input("Average change in net general government debt (% of GDP)", value=3.0, step=0.1, key="fis_perf_change_net_debt_gdp")
        with c2:
            overlap_trend = st.selectbox("Se cair em faixa sobreposta da Table 5, qual a tendência?", ["estável", "melhorando", "piorando"], index=0, key="fis_perf_overlap_trend")
        perf_init = table5_initial_from_inputs(change_net_debt_gdp, overlap_trend)
        st.metric("Initial assessment (Table 5)", perf_init)
        with st.expander("Ajustes positivos (melhoram 1 categoria cada) — Table 5", expanded=False):
            fis_perf_pos1 = st.checkbox("Government with large liquid financial assets", key="fis_perf_pos_liquid_assets")
            fis_perf_pos2 = st.checkbox("Greater ability to raise revenues / cut expenditures in the short term", key="fis_perf_pos_flexibility")
        with st.expander("Ajustes negativos (pioram 1 categoria cada) — Table 5", expanded=False):
            fis_perf_neg1 = st.checkbox("Unsustainable or volatile revenue base", key="fis_perf_neg_volatile_revenue")
            fis_perf_neg2 = st.checkbox("Limited ability to raise revenues in the short term", key="fis_perf_neg_limited_revenue")
            fis_perf_neg3 = st.checkbox("Shortfalls in basic services and infrastructure", key="fis_perf_neg_infra")
            fis_perf_neg4 = st.checkbox("Unaddressed medium-term age-related expenditure pressure", key="fis_perf_neg_ageing")
        perf_pos_count = int(fis_perf_pos1) + int(fis_perf_pos2)
        perf_neg_count = int(fis_perf_neg1) + int(fis_perf_neg2) + int(fis_perf_neg3) + int(fis_perf_neg4)
        perf_adj_raw = (-1 * perf_pos_count) + (1 * perf_neg_count)
        perf_adj = max(-2, min(2, perf_adj_raw))
        perf_final = clamp_score(perf_init + perf_adj)
        m1, m2, m3 = st.columns(3)
        m1.metric("Positivos marcados", perf_pos_count)
        m2.metric("Negativos marcados", perf_neg_count)
        m3.metric("Ajuste líquido (cap ±2)", f"{perf_adj:+d}")
        if perf_adj != perf_adj_raw:
            st.caption(f"Ajuste bruto era {perf_adj_raw:+d}, mas foi limitado para ±2 conforme Table 5.")
        st.metric("Fiscal performance & flexibility (final)", perf_final)
        img = ASSETS_DIR / "page_31_img_01.png"
        with st.expander("Ver Tabela 5 (imagem do PDF)"):
            if img.exists(): show_image(img)
            else: st.info("Imagem da Tabela 5 não encontrada em assets/.")
        st.markdown("---")
        st.markdown("## 2) Debt burden (Table 6)")
        c1, c2 = st.columns(2)
        with c1:
            net_debt_gdp = st.number_input("Net general government debt (% of GDP)", value=65.0, step=0.1, key="fis_debt_net_debt_gdp")
        with c2:
            interest_to_rev = st.number_input("General government interest expenditures (% of revenues)", value=8.0, step=0.1, key="fis_debt_interest_to_rev")
        debt_init = table6_initial_from_inputs(net_debt_gdp, interest_to_rev)
        st.metric("Initial assessment (Table 6)", debt_init)
        with st.expander("Ajuste positivo (melhora 1 categoria) — Table 6", expanded=False):
            debt_pos1 = st.checkbox("Official concessional financing likely covers gross borrowing requirements in the next 2–3 years", key="fis_debt_pos_concessional")
        with st.expander("Ajuste negativo por debt structure / funding access (piora 1 categoria se 2+ condições) — Table 6", expanded=False):
            debt_neg_cond1 = st.checkbox("More than 40% of gross debt is in foreign currency OR average maturity is typically below 3 years", key="fis_debt_neg_fx_or_maturity")
            debt_neg_cond2 = st.checkbox("Nonresidents hold consistently more than 60% of government commercial debt", key="fis_debt_neg_nonresidents")
            debt_neg_cond3 = st.checkbox("Debt service profile is subject to significant variations", key="fis_debt_neg_lumpy_profile")
            debt_neg_cond4 = st.checkbox("Banking sector exposure to government is typically above 20% of assets", key="fis_debt_neg_bank_exposure")
        debt_neg_conditions_count = int(debt_neg_cond1) + int(debt_neg_cond2) + int(debt_neg_cond3) + int(debt_neg_cond4)
        debt_structure_adj = 1 if (net_debt_gdp > 10 and debt_neg_conditions_count >= 2) else 0
        st.markdown("### Contingent liabilities (Table 7, aplicadas dentro da Table 6)")
        c1, c2 = st.columns(2)
        with c1:
            bicra = st.selectbox("BICRA group", ["1-5", "6-7", "8-9", "10"], key="fis_bicra_group")
        with c2:
            bank_assets = st.selectbox("Banks' assets / GDP", ["<=50%", "50-100%", "100-250%", "250-500%", ">500%"], key="fis_bank_assets_bucket")
        cl_cat = CONTINGENT_TABLE7[bicra][bank_assets]
        st.write(f"Categoria sugerida (Table 7): **{cl_cat}**")
        if CONTINGENT_TO_DEBT_ADJ.get(cl_cat) is None:
            chosen = st.selectbox("Escolha a categoria final para contingent liabilities", [c.strip() for c in cl_cat.split("/")], key="fis_cl_choose")
            cl_adj = CONTINGENT_TO_DEBT_ADJ[chosen]
        else:
            cl_adj = CONTINGENT_TO_DEBT_ADJ[cl_cat]
        debt_pos_adj = -1 if debt_pos1 else 0
        debt_adj_raw = debt_pos_adj + debt_structure_adj + int(cl_adj)
        debt_adj = max(-1, min(3, debt_adj_raw))
        debt_final = clamp_score(debt_init + debt_adj)
        d1, d2, d3, d4 = st.columns(4)
        d1.metric("Condições negativas marcadas", debt_neg_conditions_count)
        d2.metric("Ajuste por estrutura/funding", f"{debt_structure_adj:+d}")
        d3.metric("Ajuste por contingent liabilities", f"{int(cl_adj):+d}")
        d4.metric("Ajuste líquido (cap -1 / +3)", f"{debt_adj:+d}")
        if debt_adj != debt_adj_raw:
            st.caption(f"Ajuste bruto era {debt_adj_raw:+d}, mas foi limitado conforme Table 6.")
        st.metric("Debt burden (final)", debt_final)
        img = ASSETS_DIR / "page_36_img_01.png"
        with st.expander("Ver Tabela 6 (imagem do PDF)"):
            if img.exists(): show_image(img)
            else: st.info("Imagem da Tabela 6 não encontrada em assets/.")
        st.markdown("---")
        fiscal_final = round_to_half((perf_final + debt_final) / 2.0)
        st.metric("Fiscal assessment (average of the two segments)", fmt_score(fiscal_final))
        st.session_state["fiscal"] = float(fiscal_final)

    elif method_page == "Monetary":
        st.title("Monetary assessment")
        st.caption(
            "Cálculo baseado nas Tables 8A e 8B, com combinação de 40% para exchange-rate regime "
            "e 60% para monetary policy credibility, além dos ajustes negativos previstos na metodologia."
        )
    
        st.markdown("## 1) Exchange-rate regime")
        c1, c2 = st.columns([1.2, 0.8])
    
        with c1:
            exr_score = st.selectbox(
                "Exchange-rate regime – initial assessment",
                options=[row["Score"] for row in MONETARY_TABLE8A],
                index=1,
                key="mon_exr_score",
                format_func=lambda x: next(
                    row["Exchange-rate regime"]
                    for row in MONETARY_TABLE8A
                    if row["Score"] == x
                ),
            )
        with c2:
            st.metric("Exchange-rate regime score", exr_score)
    
        img = ASSETS_DIR / "page_27_img_01.png"
        with st.expander("Ver Tabela 8A (imagem do PDF)", expanded=False):
            if img.exists():
                show_image(img)
            else:
                st.info("Imagem da Tabela 8A não encontrada em assets/.")
    
        st.markdown("---")
        st.markdown("## 2) Monetary policy credibility")
    
        def build_mon_cred_option_blocks(score: int):
            crit = MONETARY_TABLE8B[score]
            intro = (
                "All or most of the following factors apply"
                if score in [1, 2, 3, 4]
                else "Any of the following factors apply"
            )
    
            factors = [
                crit["monetary_authority_independence"],
                crit["monetary_policy_tools_and_effectiveness"],
                crit["price_stability"],
                crit["lender_of_last_resort"],
                crit["local_financial_system_and_capital_markets"],
            ]
            factors = [f for f in factors if f and str(f).strip()]
            return intro, factors
    
        cred_score = st.selectbox(
            "Escolha o nível de monetary policy credibility conforme Table 8B",
            [MONETARY_TABLE8B_SUMMARY[i] for i in [1, 2, 3, 4, 5, 6]],
            index=2,
            key="mon_cred_choice",
        )
    
        cred_score = int(str(cred_score).split("–")[0].strip())
    
        with st.expander("Ver fatores do nível selecionado", expanded=False):
            intro, factors = build_mon_cred_option_blocks(cred_score)
            st.markdown(f"**{intro}**")
            for factor in factors:
                st.markdown(f"- {factor}")
    
        c1, c2 = st.columns([1.55, 0.45])
        with c1:
            img = ASSETS_DIR / "page_28_img_01.png"
            with st.expander("Ver Tabela 8B (imagem do PDF)", expanded=False):
                if img.exists():
                    show_image(img)
                else:
                    st.info("Imagem da Tabela 8B não encontrada em assets/.")
        with c2:
            st.metric("Monetary policy credibility score", cred_score)
    
        initial_monetary = 0.4 * float(exr_score) + 0.6 * float(cred_score)
    
        st.markdown("### Initial monetary assessment")
        m1, m2, m3 = st.columns(3)
        m1.metric("Exchange-rate regime (40%)", exr_score)
        m2.metric("Credibility (60%)", cred_score)
        m3.metric("Initial assessment", f"{initial_monetary:.1f}")
    
        st.markdown("---")
        st.markdown("## 3) Negative adjustments")
        with st.expander(
            "Ajustes negativos aplicáveis ao sovereign (máximo de 2 categorias)",
            expanded=False,
        ):
            mon_neg_1 = st.checkbox(
                "Weak or significantly weakening transmission mechanisms",
                key="mon_neg_transmission",
            )
            mon_neg_2 = st.checkbox(
                "Dollarization: resident deposits or loans in foreign currency exceed roughly 50% of total",
                key="mon_neg_dollarization",
            )
            mon_neg_3 = st.checkbox(
                "Extensive exchange restrictions (for example, IMF Article VIII issues)",
                key="mon_neg_exchange_restrictions",
            )
    
        base_neg_raw = int(mon_neg_1) + int(mon_neg_2) + int(mon_neg_3)
        base_neg = min(2, base_neg_raw)
    
        st.markdown("---")
        st.markdown("## 4) Sovereigns in monetary unions")
        in_monetary_union = st.checkbox(
            "Sovereign is part of a monetary union",
            key="mon_in_monetary_union",
        )
    
        union_adj_raw = 0
        union_adj = 0
    
        if in_monetary_union:
            dominant_member = st.checkbox(
                "Economy accounts for more than 50% of the monetary union GDP (do not apply the two union-specific adjustments)",
                key="mon_union_dominant_member",
            )
            if not dominant_member:
                with st.expander(
                    "Ajustes específicos de membros de monetary union (máximo de 2 categorias)",
                    expanded=False,
                ):
                    mu_neg_1 = st.checkbox(
                        "Member states generally have less flexibility than sovereigns with their own central bank",
                        key="mon_union_less_flexibility",
                    )
                    mu_neg_2 = st.checkbox(
                        "Economy is unsynchronized with the monetary union / the union stance may be inappropriate for this sovereign",
                        key="mon_union_unsynchronized",
                    )
                union_adj_raw = int(mu_neg_1) + int(mu_neg_2)
                union_adj = min(2, union_adj_raw)
            else:
                st.info(
                    "Como o país representa mais de 50% do PIB da união monetária, os dois ajustes específicos da união monetária não são aplicados."
                )
    
        st.markdown("---")
        final_monetary = min(6.0, initial_monetary + float(base_neg) + float(union_adj))
    
        st.markdown("## 5) Final monetary assessment")
        f1, f2, f3, f4 = st.columns(4)
        f1.metric("Initial assessment", f"{initial_monetary:.1f}")
        f2.metric("Negative adjustments", f"+{base_neg}")
        f3.metric("Monetary-union adjustments", f"+{union_adj}")
        f4.metric("Monetary assessment (final)", fmt_score(final_monetary))
    
        st.session_state["monetary"] = float(final_monetary)

    elif method_page == "External":
        st.title("External assessment")
        st.caption("Indicadores-chave e ajustes qualitativos conforme a Tabela 4 (ajuste máximo de ±3 categorias).")
        col1, col2, col3 = st.columns(3)
        with col1:
            car = st.number_input("CAR (US$) – current account receipts", min_value=0.0, value=100.0, step=1.0, key="ext_car")
            cap = st.number_input("CAP (US$) – current account payments", min_value=0.0, value=100.0, step=1.0, key="ext_cap")
        with col2:
            usable_res = st.number_input("Usable reserves (US$)", min_value=0.0, value=50.0, step=1.0, key="ext_res")
            short_term_debt = st.number_input("Short-term external debt (US$)", min_value=0.0, value=30.0, step=1.0, key="ext_st_debt")
        with col3:
            lt_maturing = st.number_input("LT external debt maturing within year (US$)", min_value=0.0, value=10.0, step=1.0, key="ext_lt_mat")
            net_ext_debt = st.number_input("Narrow net external debt (US$)", value=0.0, step=1.0, key="ext_net")
        gross_fin_needs = cap + short_term_debt + lt_maturing
        denom = car + usable_res
        liquidity_ratio = (gross_fin_needs / denom * 100.0) if denom > 0 else None
        ratio_car = (net_ext_debt / car * 100.0) if car > 0 else None
        ratio_cap = (net_ext_debt / cap * 100.0) if cap > 0 else None
        st.markdown("---")
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Gross external financing needs", f"{gross_fin_needs:,.2f}")
        m2.metric("Liquidity ratio = GEFN / (CAR+res)", "—" if liquidity_ratio is None else f"{liquidity_ratio:,.1f}%")
        m3.metric("Narrow net external debt / CAR", "—" if ratio_car is None else f"{ratio_car:,.1f}%")
        m4.metric("Narrow net external debt / CAP", "—" if ratio_cap is None else f"{ratio_cap:,.1f}%")
        st.markdown("---")
        init_ext = st.selectbox("External assessment – score inicial (1–6)", [1, 2, 3, 4, 5, 6], index=2, key="ext_score")
        with st.expander("Ajustes positivos (melhoram 1 categoria cada)", expanded=False):
            pos1 = st.checkbox("País com posição líquida externa significativamente mais forte", key="ext_pos_net_position")
            pos2 = st.checkbox("Moeda ativamente negociada + superávits recorrentes em conta corrente", key="ext_pos_atc_surplus")
        with st.expander("Ajustes negativos (pioram 1 categoria cada; exceto quando indicado)", expanded=False):
            neg1 = st.checkbox("Risco de deterioração acentuada no financiamento externo", key="ext_neg_financing")
            neg2 = st.checkbox("Volatilidade significativa nos termos de troca", key="ext_neg_tot")
            neg3 = st.checkbox("Baixa dívida externa reflete restrições de endividamento", key="ext_neg_constraints")
            neg4 = st.checkbox("Inconsistências materiais de dados", key="ext_neg_data")
            neg5 = st.checkbox("Moeda ativamente negociada + altos déficits em conta corrente", key="ext_neg_atc_high_def")
            neg6 = st.checkbox("(piora 2 categorias) Moeda ativamente negociada + déficits muito altos em conta corrente", key="ext_neg_atc_very_high_def")
        pos_count = int(pos1) + int(pos2)
        neg_count = int(neg1) + int(neg2) + int(neg3) + int(neg4) + int(neg5) + int(neg6)
        pos_adj = -1 * int(pos1) + -1 * int(pos2)
        neg_adj = 1 * int(neg1) + 1 * int(neg2) + 1 * int(neg3) + 1 * int(neg4) + 1 * int(neg5) + 2 * int(neg6)
        total_adj_raw = pos_adj + neg_adj
        total_adj = max(-3, min(3, total_adj_raw))
        s1, s2, s3 = st.columns([1, 1, 1])
        s1.metric("Positivos marcados", pos_count)
        s2.metric("Negativos marcados", neg_count)
        s3.metric("Ajuste líquido (cap ±3)", f"{total_adj:+d}")
        if total_adj != total_adj_raw:
            st.caption(f"Ajuste bruto era {total_adj_raw:+d}, mas foi limitado para ±3 conforme Table 4.")
        final_ext = clamp_score(init_ext + total_adj)
        st.metric("External assessment (final)", final_ext)
        img = ASSETS_DIR / "page_25_img_01.png"
        with st.expander("Ver Tabela 4 (imagem do PDF)"):
            if img.exists(): show_image(img)
            else: st.info("Imagem da Tabela 4 não encontrada em assets/.")
        st.session_state["external"] = int(final_ext)

    elif method_page == "Institutional":
        st.title("Institutional assessment")
        st.markdown("#### Seleção do nível (Tabela 2)")
        inst_choice = st.selectbox("Escolha o nível institucional conforme Table 2", [INST_TABLE2_LABELS[i] for i in [1, 2, 3, 4, 5, 6]], index=3, key="inst_table2_choice")
        init_inst = int(inst_choice.split("–")[0].strip())
        with st.expander("Ver critérios do Table 2 (nível selecionado)", expanded=False):
            st.markdown("**Effectiveness, stability, and predictability of policymaking, political institutions, and civil society**")
            st.markdown(bullets(INST_TABLE2[init_inst]["effectiveness"]))
            st.markdown("**Transparency and accountability of institutions, data, and processes**")
            st.markdown(bullets(INST_TABLE2[init_inst]["transparency"]))
        st.markdown("#### Ajustes")
        debt_culture_risk = st.checkbox("Risco de debt payment culture (cap para 6)", value=False, key="inst_debt_culture")
        war_risk = st.selectbox("External security risk (se aplicável)", options=[0, 1, 2], index=0, key="inst_war_adj")
        if debt_culture_risk:
            final_inst = 6
        else:
            final_inst = clamp_score(init_inst + war_risk)
        st.metric("Institutional assessment (final)", final_inst)
        st.session_state["institutional"] = int(final_inst)

    elif method_page == "Resultados":
        st.title("Resultados e perfis")
        institutional = int(st.session_state.get("institutional", 4))
        economic = int(st.session_state.get("economic", 4))
        external = int(st.session_state.get("external", 3))
        fiscal = float(st.session_state.get("fiscal", 4))
        monetary = float(st.session_state.get("monetary", 3))
        ie_profile = (institutional + economic) / 2.0
        fp_profile = (external + fiscal + monetary) / 3.0
        profiles = {
            "Institutional & Economic profile": round(ie_profile, 2),
            "Flexibility & Performance profile": round(fp_profile, 2),
        }
        st.session_state["profiles"] = profiles
        c1, c2, c3 = st.columns([1, 1, 1])
        c1.metric("Institutional", institutional)
        c2.metric("Economic", economic)
        c3.metric("Institutional & Economic profile (avg)", f"{ie_profile:.2f}")
        c1, c2, c3, c4 = st.columns([1, 1, 1, 1])
        c1.metric("External", external)
        c2.metric("Fiscal", fmt_score(fiscal))
        c3.metric("Monetary", fmt_score(monetary))
        c4.metric("Flexibility & Performance profile (avg)", f"{fp_profile:.2f}")
        st.markdown("---")
        st.subheader("Radar (1=melhor; 6=pior)")
        fig = radar({"Institutional": institutional, "Economic": economic, "External": external, "Fiscal": fiscal, "Monetary": monetary})
        st_plotly_chart_compat(fig, use_container_width=True)
        st.markdown("---")
        st.subheader("Indicative rating level & notches")
        img = ASSETS_DIR / "page_06_img_01.png"
        if img.exists():
            show_image(img)
        indicative_upper = indicative_from_matrix(ie_profile, fp_profile)
        indicative_lower = indicative_upper.lower()
        st.metric("Indicative rating level", indicative_lower)
        with st.expander("Override manual (opcional)"):
            st.selectbox("Indicative rating level (manual override)", RATING_SCALE, index=RATING_SCALE.index(indicative_upper), key="indicative_level_override")
        base_rating = st.session_state.get("indicative_level_override", indicative_upper)
        notch_adj = st.selectbox("Ajuste de notches (−1, 0, +1)", [-1, 0, 1], index=1, key="notch_adj")
        lc_uplift = st.selectbox("Uplift para Local-Currency (0 ou +1)", [0, 1], index=0, key="lc_uplift")
        final_rating = apply_notches(base_rating, notch_adj, lc_uplift)
        st.session_state["indicative"] = indicative_lower
        st.session_state["final_rating"] = final_rating
        st.metric("Final rating", final_rating)
        st.write(f"Notch: **{notch_adj:+d}** LC uplift: **+{lc_uplift}**")


# ============================================================
# App principal
# ============================================================

def main():
    st.title("S&P Methodology + SRI")
    st.caption("1ª aba = metodologia; abas seguintes = dashboards do SRI")

    local_file = find_local_xlsx()
    with st.expander("Arquivo de entrada do SRI", expanded=False):
        uploaded = st.file_uploader("Se quiser, envie um arquivo .xlsx para substituir a base local", type=["xlsx"])
        if uploaded is None and local_file is not None:
            try:
                rel = local_file.relative_to(APP_DIR)
            except Exception:
                rel = local_file
            st.success(f"Usando arquivo do repositório: {rel}")
        elif uploaded is None:
            st.info("Nenhum arquivo local encontrado. Faça upload de um .xlsx ou adicione um arquivo em ./data.")

    uploaded_bytes = uploaded.getvalue() if uploaded is not None else None
    df = load_workbook(uploaded_bytes)
    filtered = None
    if not df.empty:
        with st.expander("Filtros do SRI", expanded=False):
            filtered = build_filters(df)

    tab1, tab2, tab3 = st.tabs(["Metodologia", "SRI – Dashboards", "SRI – Dados em tabela"])

    with tab1:
        render_methodology_tab()

    with tab2:
        if df.empty or filtered is None:
            st.error("Não foi possível interpretar a estrutura do workbook.")
        else:
            render_dashboard_tab(filtered)

    with tab3:
        if df.empty or filtered is None:
            st.error("Não foi possível interpretar a estrutura do workbook.")
        else:
            render_table_tab(filtered)
            with st.expander("Dicionário de campos", expanded=False):
                st.markdown(
                    """
- **sheet**: nome da aba original da planilha.
- **country_name**: nome do país.
- **country_code**: código do país.
- **lt_fc_rating**: rating LT FC.
- **indicator**: nome do indicador.
- **year**: ano original da base (preserva `e` e `f`).
- **year_num**: ano numérico para ordenação.
- **is_forecast**: identifica estimativa/projeção.
- **value**: valor numérico convertido para análise.
                    """
                )

if __name__ == "__main__":
    main()
